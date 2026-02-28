import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait

def update_excel_data(filepath, searchTerm, colName,new_value):
    book = openpyxl.load_workbook(filepath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row = 1, column = i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row = i, column = j).value == searchTerm:
                Dict["row"] = i


#Code for Editing the excel with updated value
    sheet.cell(row = Dict["row"], column = Dict["col"]).value = new_value
    book.save(file_path)


file_path = "C:\\Users\\Dell\\Downloads\\download.xlsx"
fruit_name = "Mango"
new_value = "999"
options = Options()
options.add_experimental_option("detach", True)
driver = webdriver.Chrome(options=options)
driver.maximize_window()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

#edit the excel with updated values
update_excel_data(file_path, fruit_name, "price",new_value)

#upload
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 10)
wait.until(expected_conditions.visibility_of_element_located((By.XPATH, "//div[text()='Updated Excel Data Successfully.']")))
print(driver.find_element(By.XPATH, "//div[text()='Updated Excel Data Successfully.']").text)
#pricecolumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_price = driver.find_element(By.XPATH,"//div[text()='Mango']/parent::div/parent::div/div[@id='cell-4-undefined']").text
assert actual_price == new_value
