import openpyxl

book = openpyxl.load_workbook("D:\\Automation.xlsx")
sheet = book.active
Dict = {}
#print(sheet.cell(row = 1, column = 2).value)
for i in range(1, sheet.max_row + 1): # To get rows
    if sheet.cell(row = i, column = 1).value == "Testcase2":
        for j in range(2, sheet.max_column+1): # TO get columns
            Dict[sheet.cell(row = 1, column = j).value] = sheet.cell(row = i, column = j).value
print(Dict)